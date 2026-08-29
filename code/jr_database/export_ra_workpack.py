#!/usr/bin/env python3
"""Build the single RA workbook: resolve + Sheet2 reciprocity + ICMID import.

Writes only: output/jr_database/RA_workpack.xlsx

Usage (from ICMID PingJu project root):
    uv run python -B code/jr_database/export_ra_workpack.py

Typical loop:
    uv run python -B code/jr_database/build_cross_group.py --apply-unmatched
    uv run python -B code/jr_database/export_ra_workpack.py
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_CODE = Path(__file__).resolve().parent.parent
_PIPELINE = _CODE.parent
for _p in (_CODE, _PIPELINE, _CODE / "visualization"):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

from jr_database.build_cross_group import build_pair_table, build_unmatched  # noqa: E402
from jr_database.lib.sync_icmid_unmatched import sync_icmid_unmatched_sheet  # noqa: E402
from jr_database.config import (  # noqa: E402
    ICMID_MANUAL_XLSX,
    RA_WORKPACK_XLSX,
    ensure_output_dirs,
)
from jr_database.lib.sheet_worklist import build_worklist  # noqa: E402
from jr_database.lib.one_sided import ResolveCache, build_tables, load_sheet2  # noqa: E402
from jr_database.resolve_homeland import get_resolver  # noqa: E402
from jr_database.sources import load_all_cross_assertions  # noqa: E402


def _steps_df() -> pd.DataFrame:
    rows = [
        (
            "Purpose",
            "One workbook for three tasks: resolve names, fix Sheet2 reciprocity gaps, "
            "and import Keerthana/llm pairs into ICMID Sheet2/3.",
        ),
        ("", ""),
        (
            "Principle",
            "Resolve names first, then judge missing / one-sided links. "
            "Unresolved spellings are not missing reverse links.",
        ),
        (
            "Resolve order",
            "ethnic_entity_index → registry aliases → GIS exact (murdock → greg → geopr).",
        ),
        (
            "Allowed polygon_source",
            "Only murdock / greg / geopr (do not use Joshua Project).",
        ),
        ("", ""),
        ("STEP 1 — Resolve names (required)", "Open sheet: 1_unmatched_entities"),
        (
            "What to do",
            "Entities still without a homeland. Fill polygon_source, polygon_id; "
            "optional display_name / resolve_source / aliases.",
        ),
        (
            "Example",
            "Maasai → murdock / MASAI (double-a spelling must be mapped to Murdock MASAI).",
        ),
        (
            "After filling",
            "Save, then run: uv run python -B code/jr_database/build_cross_group.py --apply-unmatched",
        ),
        ("", ""),
        (
            "STEP 1b — Unresolved Sheet2 partners",
            "Open sheet: 1b_unresolved_partners",
        ),
        (
            "What to do",
            "Spellings in Sheet2 Joking link that do not resolve (e.g. Zaramu, Kami). "
            "Add them to 1_unmatched_entities or an index alias, then --apply-unmatched.",
        ),
        (
            "Note",
            "These are not missing reverse links; alias first, then re-export the workpack "
            "and check 2_missing_reverse.",
        ),
        ("", ""),
        ("STEP 2 — Sheet2 missing reverse", "Open sheet: 2_missing_reverse"),
        (
            "What to do",
            "has_link already lists the partner; missing_on does not list back. "
            "Add should_add on ICMID- Africa.xlsx Sheet2, or remove from has_link and note why.",
        ),
        (
            "If unresolved spellings remain",
            "Finish STEP 1b first, then re-export to avoid false positives.",
        ),
        ("", ""),
        (
            "STEP 2b — Partner has no Sheet2 row",
            "Open sheet: 2b_partner_not_in_sheet2",
        ),
        (
            "Meaning",
            "A lists B, B resolves, but Sheet2 has no row for B (often GREG/GeoEPR). "
            "Cannot add a Sheet2 reverse link.",
        ),
        (
            "How to handle",
            "Accept as an extra group, or confirm whether a Murdock row should be created; "
            "do not treat as missing_reverse.",
        ),
        ("", ""),
        (
            "STEP 3 — Import other sources into ICMID",
            "Open sheets: 3_to_sheet2 and 3_to_sheet3",
        ),
        (
            "Rule (advisor)",
            "If at least one endpoint is murdock → Sheet2; if neither is → Sheet3.",
        ),
        (
            "Sources",
            "keerthana_analysis + llm_ehraf (not keerthana og).",
        ),
        (
            "3_to_sheet2",
            "status=add_partner… → append partner_to_add on the existing Sheet2 Joking link; "
            "create_new… → create an Ethnic Group row first.",
        ),
        (
            "3_to_sheet3",
            "Neither side is murdock; candidate for Sheet3. Drop overly generic names "
            "(villagers, Europeans, males, …) first.",
        ),
        ("3_already_on_sheet2", "Already on Sheet2; skip."),
        ("", ""),
        (
            "Suggested order",
            "1 → 1b → apply-unmatched → re-export this workpack → 2 → 2b → 3.",
        ),
        (
            "Re-run commands",
            "uv run python -B code/jr_database/build_cross_group.py --apply-unmatched && "
            "uv run python -B code/jr_database/export_ra_workpack.py",
        ),
        (
            "Primary coding file",
            "data/sources/ICMID- Africa.xlsx (Sheet2 main coding; Sheet3 holding).",
        ),
        (
            "Long-term lookup",
            "data/sources/ICMID- Africa.xlsx → sheets ethnic_entity_index + Unmatched entities "
            "(synced on build / export_ra_workpack).",
        ),
    ]
    return pd.DataFrame(rows, columns=["section", "detail"])


def build_pack(src: Path = ICMID_MANUAL_XLSX) -> dict[str, pd.DataFrame]:
    print("Loading assertions + unmatched…")
    assertions = load_all_cross_assertions()
    pairs = build_pair_table(assertions)
    unmatched = build_unmatched(pairs, assertions)
    print(f"  unmatched={len(unmatched)}")

    kept, removed, added = sync_icmid_unmatched_sheet(unmatched)
    print(f"  synced ICMID Unmatched entities: {kept} rows (-{removed} resolved, +{added} new)")

    print(f"Sheet2 one-sided from {src.name}…")
    sheet2 = load_sheet2(src)
    one = build_tables(sheet2, ResolveCache(resolver=get_resolver()))
    missing = one["missing_reverse"]
    unresolved = one["unresolved_partners"]
    no_row = one["partner_not_in_sheet2"]
    one_sum = one["summary"]
    print(
        f"  missing_reverse={len(missing)}  "
        f"unresolved={len(unresolved)}  "
        f"partner_not_in_sheet2={len(no_row)}"
    )

    print("ICMID Sheet2/3 import worklist…")
    wl = build_worklist(src)
    to_s2 = wl["to_sheet2"]
    to_s3 = wl["to_sheet3"]
    already = wl["already_on_sheet2"]
    wl_sum = wl["summary"]
    print(
        f"  to_sheet2={len(to_s2)}  to_sheet3={len(to_s3)}  already={len(already)}"
    )

    counts = pd.DataFrame(
        [
            {"sheet": "1_unmatched_entities", "n": len(unmatched), "step": "1 resolve names"},
            {"sheet": "1b_unresolved_partners", "n": len(unresolved), "step": "1b resolve Sheet2 spellings"},
            {"sheet": "2_missing_reverse", "n": len(missing), "step": "2 fix Sheet2 reciprocity"},
            {"sheet": "2b_partner_not_in_sheet2", "n": len(no_row), "step": "2b no Sheet2 row"},
            {
                "sheet": "3_to_sheet2",
                "n": len(to_s2),
                "step": "3 import ≥1 murdock side into Sheet2",
            },
            {
                "sheet": "3_to_sheet3",
                "n": len(to_s3),
                "step": "3 import neither-murdock into Sheet3",
            },
            {
                "sheet": "3_already_on_sheet2",
                "n": len(already),
                "step": "3 skip (already coded)",
            },
        ]
    )

    combined_summary_rows = []
    if not one_sum.empty:
        for _, r in one_sum.iterrows():
            combined_summary_rows.append(
                {"block": "Sheet2 one-sided", "item": r.get("item", ""), "n": r.get("n", "")}
            )
    if not wl_sum.empty:
        for _, r in wl_sum.iterrows():
            combined_summary_rows.append(
                {"block": "ICMID import", "item": r.get("item", ""), "n": r.get("n", "")}
            )
    combined_summary_rows.append(
        {"block": "unmatched", "item": "unmatched entities", "n": len(unmatched)}
    )
    combined_summary = pd.DataFrame(combined_summary_rows)

    return {
        "0_STEPS": _steps_df(),
        "0_counts": counts,
        "0_summary": combined_summary,
        "1_unmatched_entities": unmatched,
        "1b_unresolved_partners": unresolved,
        "2_missing_reverse": missing,
        "2b_partner_not_in_sheet2": no_row,
        "3_to_sheet2": to_s2,
        "3_to_sheet3": to_s3,
        "3_already_on_sheet2": already,
    }


def _style(ws) -> None:
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    headers = [c.value for c in ws[1]]
    default_w = 14
    wide = {
        "section": 28,
        "detail": 90,
        "action": 48,
        "quote": 40,
        "current_Joking_link": 36,
        "missing_on_currently_lists": 40,
        "should_add": 16,
        "entity": 18,
        "partner_raw": 20,
        "why": 48,
        "item": 48,
        "sheet": 28,
        "step": 40,
    }
    for i, name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wide.get(name, default_w)

    wrap_names = {
        "detail",
        "action",
        "quote",
        "current_Joking_link",
        "missing_on_currently_lists",
        "why",
        "step",
    }
    wrap_idx = {i for i, n in enumerate(headers, start=1) if n in wrap_names}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap if cell.column in wrap_idx else top

    if headers and headers[0] == "section":
        step_fill = PatternFill("solid", fgColor="E2EFDA")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            sec = str(row[0].value or "")
            if sec.startswith("STEP"):
                for cell in row:
                    cell.fill = step_fill


def write_xlsx(sheets: dict[str, pd.DataFrame], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            out = df if df is not None else pd.DataFrame()
            out.to_excel(writer, sheet_name=name[:31], index=False)
        for ws in writer.book.worksheets:
            _style(ws)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build single RA workpack workbook")
    parser.add_argument("--src", type=Path, default=ICMID_MANUAL_XLSX)
    parser.add_argument("--out", type=Path, default=RA_WORKPACK_XLSX)
    args = parser.parse_args()
    ensure_output_dirs()

    print("Building RA workpack…")
    sheets = build_pack(args.src)
    write_xlsx(sheets, args.out)
    print(f"  → {args.out}")
    for name, df in sheets.items():
        if name.startswith("0_"):
            continue
        print(f"  {name}: {len(df)} rows")


if __name__ == "__main__":
    main()
