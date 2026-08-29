"""Sync ICMID workbook sheet ``Unmatched entities`` with current resolve state."""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

_CODE = Path(__file__).resolve().parent.parent.parent
if str(_CODE) not in sys.path:
    sys.path.insert(0, str(_CODE))
if str(_CODE / "visualization") not in sys.path:
    sys.path.insert(0, str(_CODE / "visualization"))

from entity_index import build_lookup, load_index, lookup_row  # noqa: E402
from jr_database.config import ICMID_MANUAL_XLSX, ICMID_UNMATCHED_SHEET  # noqa: E402
from jr_database.resolve_homeland import VALID_HOMELAND, get_resolver  # noqa: E402

UNMATCHED_SHEET_COLUMNS = (
    "Name in the coding",
    "Mentions",
    "Region(s)",
    "Named by",
    "ours (all_ethnics_new)",
    "direct polygon match",
    "RA index: source",
    "RA index: polygon_id",
    "RA -> Murdock polygon",
    "in RA index at all",
    "resolves to a Murdock polygon",
    "gain from the RA file",
    "disagreement",
    "Source",
    "Notes",
)


def _clean(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


def _merge_list_field(*parts: str) -> str:
    out: list[str] = []
    seen: set[str] = set()
    for part in parts:
        for token in re_split_tokens(part):
            k = token.casefold()
            if k and k not in seen:
                seen.add(k)
                out.append(token)
    return ", ".join(out)


def re_split_tokens(text: str) -> list[str]:
    raw = _clean(text)
    if not raw:
        return []
    for sep in (";", "|", "/"):
        raw = raw.replace(sep, ",")
    return [_clean(t) for t in raw.split(",") if _clean(t)]


def _is_resolved(resolver, name: str) -> bool:
    hit = resolver.resolve(name)
    src = _clean(hit.polygon_source).lower().replace("geoepr", "geopr")
    return src in VALID_HOMELAND and bool(_clean(hit.polygon_id))


def _refresh_ra_columns(row: dict, name: str, resolver, lookup: dict) -> None:
    idx = lookup_row(lookup, name)
    hit = resolver.resolve(name)
    row["RA index: source"] = _clean(idx.get("polygon_source")) if idx is not None else ""
    row["RA index: polygon_id"] = _clean(idx.get("polygon_id")) if idx is not None else ""
    row["in RA index at all"] = "yes" if idx is not None else "no"
    src = _clean(hit.polygon_source).lower()
    row["resolves to a Murdock polygon"] = (
        "yes" if src == "murdock" and _clean(hit.polygon_id) else "NO"
    )
    if idx is not None:
        rs = _clean(idx.get("resolve_source"))
        if rs and not _clean(row.get("Source")):
            row["Source"] = rs


def _empty_row(name: str = "") -> dict:
    row = {c: None for c in UNMATCHED_SHEET_COLUMNS}
    row["Name in the coding"] = name
    return row


def _merge_existing_row(base: dict, other: dict) -> None:
    """Combine duplicate legacy rows for the same entity name."""
    try:
        m_base = float(base.get("Mentions") or 0)
        m_other = float(other.get("Mentions") or 0)
        base["Mentions"] = max(m_base, m_other) if (m_base or m_other) else None
    except (TypeError, ValueError):
        base["Mentions"] = base.get("Mentions") or other.get("Mentions")
    base["Region(s)"] = _merge_list_field(base.get("Region(s)"), other.get("Region(s)"))
    base["Named by"] = _merge_list_field(base.get("Named by"), other.get("Named by"))
    if not _clean(base.get("Notes")) and _clean(other.get("Notes")):
        base["Notes"] = other.get("Notes")
    if not _clean(base.get("Source")) and _clean(other.get("Source")):
        base["Source"] = other.get("Source")


def _write_sheet(df: pd.DataFrame) -> None:
    wb = load_workbook(ICMID_MANUAL_XLSX)
    if ICMID_UNMATCHED_SHEET in wb.sheetnames:
        del wb[ICMID_UNMATCHED_SHEET]
    ws = wb.create_sheet(ICMID_UNMATCHED_SHEET)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    wb.save(ICMID_MANUAL_XLSX)


def sync_icmid_unmatched_sheet(unmatched: pd.DataFrame) -> tuple[int, int, int]:
    """Drop resolved names; one row per entity; refresh RA columns.

    Returns (kept_rows, removed_resolved, added_new).
    """
    if not ICMID_MANUAL_XLSX.is_file():
        return 0, 0, 0

    get_resolver.cache_clear()
    resolver = get_resolver()
    lookup = build_lookup(load_index())

    try:
        existing = pd.read_excel(ICMID_MANUAL_XLSX, sheet_name=ICMID_UNMATCHED_SHEET)
    except ValueError:
        existing = pd.DataFrame(columns=list(UNMATCHED_SHEET_COLUMNS))

    um_by_key = {
        _clean(r.get("entity")).casefold(): r
        for _, r in unmatched.iterrows()
        if _clean(r.get("entity"))
    }

    existing_by_key: dict[str, dict] = {}
    removed = 0

    for _, r in existing.iterrows():
        name = _clean(r.get("Name in the coding"))
        if not name:
            continue
        if _is_resolved(resolver, name):
            removed += 1
            continue
        key = name.casefold()
        row = {c: r.get(c, None) for c in UNMATCHED_SHEET_COLUMNS if c in r.index or c == "Name in the coding"}
        if "Name in the coding" not in row or row["Name in the coding"] is None:
            row["Name in the coding"] = name
        for c in UNMATCHED_SHEET_COLUMNS:
            row.setdefault(c, None)
        if key in existing_by_key:
            _merge_existing_row(existing_by_key[key], row)
        else:
            existing_by_key[key] = row

    added = 0
    all_keys = sorted(set(existing_by_key) | set(um_by_key))
    kept: list[dict] = []

    for key in all_keys:
        row = existing_by_key.get(key, _empty_row())
        if key in um_by_key:
            um = um_by_key[key]
            if key not in existing_by_key:
                added += 1
            row["Name in the coding"] = _clean(um.get("entity")) or row.get("Name in the coding")
            row["Mentions"] = um.get("n_pairs")
            row["Region(s)"] = um.get("region")
            row["Named by"] = _merge_list_field(row.get("Named by"), um.get("example_pair_partner"))
            if _clean(um.get("notes")):
                row["Notes"] = um.get("notes")
            if _clean(um.get("resolve_source")):
                row["Source"] = um.get("resolve_source")
        name = _clean(row.get("Name in the coding"))
        if not name:
            continue
        _refresh_ra_columns(row, name, resolver, lookup)
        kept.append(row)

    out = pd.DataFrame(kept)
    for c in UNMATCHED_SHEET_COLUMNS:
        if c not in out.columns:
            out[c] = None
    out = out[list(UNMATCHED_SHEET_COLUMNS)]
    if not out.empty:
        out = out.sort_values(
            ["Mentions", "Name in the coding"],
            ascending=[False, True],
            kind="mergesort",
            na_position="last",
        )
    _write_sheet(out)
    return len(out), removed, added
