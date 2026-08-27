#!/usr/bin/env python3
"""Sheet2 one-sided JR helpers for export_ra_workpack (internal).

Do not run directly — use export_ra_workpack.py.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_JR = Path(__file__).resolve().parent.parent  # code/jr_database
_CODE = _JR.parent
_PIPELINE = _CODE.parent
for _p in (_CODE, _PIPELINE, _CODE / "visualization"):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

from jr_database.config import (  # noqa: E402
    ICMID_MANUAL_XLSX,
    ensure_output_dirs,
)
from jr_database.resolve_homeland import VALID_HOMELAND, HomelandHit, get_resolver  # noqa: E402
from jr_database.sources import _SKIP_JOKING_LINK, _SKIP_PARTNER_TOKEN, _clean  # noqa: E402

_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_AND_PREFIX_RE = re.compile(r"^(and|&)\s+", re.I)
_PAREN_RE = re.compile(r"\((.+?)\)")


def _excel_safe(val: Any) -> Any:
    if isinstance(val, str):
        return _ILLEGAL_XLSX_RE.sub("", val)
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return val


def parse_joking_partners(val: Any) -> list[str]:
    s = _clean(val)
    if not s or s.casefold() in _SKIP_JOKING_LINK:
        return []
    low = s.casefold()
    if "http" in low or "article" in low or "access" in low:
        return []

    out: list[str] = []
    seen: set[str] = set()
    for part in re.split(r"[,;/]", s):
        p = _AND_PREFIX_RE.sub("", part.strip()).strip()
        pl = p.casefold()
        if not p or pl in _SKIP_JOKING_LINK or pl in _SKIP_PARTNER_TOKEN:
            continue
        if pl in seen:
            continue
        seen.add(pl)
        out.append(p)
    return out


def _name_candidates(name: str) -> list[str]:
    s = re.sub(r"\s+", " ", _clean(name))
    out: list[str] = []

    def add(x: str) -> None:
        x = _clean(x)
        if x and x not in out:
            out.append(x)

    add(s)
    add(re.sub(r"^the\s+", "", s, flags=re.I))
    m = _PAREN_RE.search(s)
    if m:
        add(re.sub(r"\s*\(.*?\)", "", s))
        add(m.group(1))
    for cand in list(out):
        add(cand.upper())
        add(cand.title())
    return out


def poly_key(hit: HomelandHit) -> str:
    if hit.polygon_source in VALID_HOMELAND and hit.polygon_id:
        return f"{hit.polygon_source}:{hit.polygon_id.upper()}"
    return ""


def _label_hit(hit: HomelandHit) -> str:
    pid = _clean(hit.polygon_id)
    disp = _clean(hit.display_name)
    if pid.isdigit() and disp:
        return disp.upper()
    return pid or disp.upper()


@dataclass
class SheetRow:
    ethnic_group: str
    hit: HomelandHit
    partners_raw: list[str]
    region: str
    country: str
    joking_link: str

    @property
    def key(self) -> str:
        return poly_key(self.hit)


@dataclass
class ResolveCache:
    resolver: Any
    _hits: dict[str, HomelandHit] = field(default_factory=dict)

    def resolve(self, name: str) -> HomelandHit:
        raw = _clean(name)
        cache_key = raw.casefold()
        if cache_key in self._hits:
            return self._hits[cache_key]
        hit = HomelandHit("", "", raw, "", "")
        for cand in _name_candidates(raw):
            trial = self.resolver.resolve(cand)
            if trial.polygon_source in VALID_HOMELAND and trial.polygon_id:
                hit = trial
                break
        self._hits[cache_key] = hit
        return hit


def load_sheet2(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Sheet2")
    df.columns = [str(c).strip() for c in df.columns]
    if "Ethnic Group" not in df.columns:
        raise ValueError(f"Sheet2 missing 'Ethnic Group': {list(df.columns)}")
    df["Ethnic Group"] = df["Ethnic Group"].map(lambda v: re.sub(r"\s+", " ", _clean(v)))
    df = df[df["Ethnic Group"] != ""].copy()
    return df.reset_index(drop=True)


def build_tables(df: pd.DataFrame, cache: ResolveCache) -> dict[str, pd.DataFrame]:
    sheet_rows: list[SheetRow] = []
    for _, r in df.iterrows():
        name = _clean(r.get("Ethnic Group"))
        if not name:
            continue
        hit = cache.resolve(name)
        jl = r.get("Joking link") if "Joking link" in df.columns else ""
        sheet_rows.append(
            SheetRow(
                ethnic_group=name,
                hit=hit,
                partners_raw=parse_joking_partners(jl),
                region=_clean(r.get("Region")),
                country=_clean(r.get("Country_1")),
                joking_link=_excel_safe(jl) if _clean(jl) else "",
            )
        )

    by_key: dict[str, list[SheetRow]] = defaultdict(list)
    for srow in sheet_rows:
        if srow.key:
            by_key[srow.key].append(srow)

    edges: dict[tuple[str, str], list[tuple[SheetRow, str]]] = defaultdict(list)
    unresolved_rows: list[dict[str, Any]] = []
    self_links = 0

    for srow in sheet_rows:
        if not srow.key:
            for partner in srow.partners_raw:
                unresolved_rows.append(
                    {
                        "partner_raw": partner,
                        "named_by": srow.ethnic_group,
                        "named_by_polygon": "",
                        "Region": srow.region,
                        "Country_1": srow.country,
                        "why": "source Ethnic Group unresolved",
                    }
                )
            continue
        for partner in srow.partners_raw:
            phit = cache.resolve(partner)
            b_key = poly_key(phit)
            if not b_key:
                unresolved_rows.append(
                    {
                        "partner_raw": partner,
                        "named_by": srow.ethnic_group,
                        "named_by_polygon": srow.hit.polygon_id,
                        "Region": srow.region,
                        "Country_1": srow.country,
                        "why": "partner spelling did not resolve — not counted as missing reverse",
                    }
                )
                continue
            if b_key == srow.key:
                self_links += 1
                continue
            edges[(srow.key, b_key)].append((srow, partner))

    outgoing: dict[str, set[str]] = defaultdict(set)
    for a_key, b_key in edges:
        outgoing[a_key].add(b_key)

    def _poly_label(key: str) -> str:
        rows = by_key.get(key, [])
        if rows:
            return _label_hit(rows[0].hit)
        return key.split(":", 1)[-1]

    missing_rows: list[dict[str, Any]] = []
    no_row_rows: list[dict[str, Any]] = []
    reciprocal_pairs: set[tuple[str, str]] = set()

    for (a_key, b_key), mentions in sorted(
        edges.items(),
        key=lambda kv: (_poly_label(kv[0][0]), _poly_label(kv[0][1])),
    ):
        if a_key in outgoing.get(b_key, set()):
            reciprocal_pairs.add(tuple(sorted((a_key, b_key))))
            continue

        a_label = _poly_label(a_key)
        b_label = _poly_label(b_key)
        named_by = "; ".join(dict.fromkeys(m[0].ethnic_group for m in mentions))
        wrote = "; ".join(dict.fromkeys(m[1] for m in mentions))
        a_region = mentions[0][0].region
        a_country = mentions[0][0].country

        if b_key not in by_key:
            phit = cache.resolve(mentions[0][1])
            no_row_rows.append(
                {
                    "has_link": named_by,
                    "wrote": wrote,
                    "partner_resolved_as": _label_hit(phit),
                    "partner_polygon_source": phit.polygon_source,
                    "partner_polygon_id": phit.polygon_id,
                    "Region": a_region,
                    "Country_1": a_country,
                    "note": "partner resolved, but no Sheet2 row to edit",
                }
            )
            continue

        b_rows = by_key[b_key]
        b_sheet_names = "; ".join(dict.fromkeys(r.ethnic_group for r in b_rows))
        b_current = " | ".join(
            f"{r.ethnic_group}: {r.joking_link or '(empty/None)'}" for r in b_rows
        )
        b_unresolved = []
        seen_u: set[str] = set()
        for brow in b_rows:
            for p in brow.partners_raw:
                if poly_key(cache.resolve(p)):
                    continue
                if p.casefold() not in seen_u:
                    seen_u.add(p.casefold())
                    b_unresolved.append(p)

        # What should be added on B's row (prefer A's Sheet2 Ethnic Group name).
        add_on_b = named_by.split("; ")[0]

        missing_rows.append(
            {
                "has_link": named_by,
                "wrote": wrote,
                "resolved_as": f"{a_label} → {b_label}",
                "missing_on": b_sheet_names,
                "should_add": add_on_b,
                "missing_on_currently_lists": b_current,
                "missing_on_unresolved_spellings": ", ".join(b_unresolved),
                "Region": a_region,
                "Country_1": a_country,
                "a_polygon_id": a_label,
                "b_polygon_id": b_label,
            }
        )

    missing_cols = [
        "has_link",
        "wrote",
        "resolved_as",
        "missing_on",
        "should_add",
        "missing_on_currently_lists",
        "missing_on_unresolved_spellings",
        "Region",
        "Country_1",
        "a_polygon_id",
        "b_polygon_id",
    ]
    missing = pd.DataFrame(missing_rows, columns=missing_cols)

    unresolved_cols = [
        "partner_raw",
        "named_by",
        "named_by_polygon",
        "Region",
        "Country_1",
        "why",
    ]
    unresolved = pd.DataFrame(unresolved_rows, columns=unresolved_cols)
    if not unresolved.empty:
        unresolved = unresolved.sort_values(
            ["partner_raw", "named_by"], kind="mergesort"
        ).reset_index(drop=True)

    no_row_cols = [
        "has_link",
        "wrote",
        "partner_resolved_as",
        "partner_polygon_source",
        "partner_polygon_id",
        "Region",
        "Country_1",
        "note",
    ]
    no_row = pd.DataFrame(no_row_rows, columns=no_row_cols)

    summary = pd.DataFrame(
        [
            {"item": "Sheet2 rows", "n": len(sheet_rows)},
            {
                "item": "Sheet2 rows with cross-group partners",
                "n": sum(1 for s in sheet_rows if s.partners_raw),
            },
            {
                "item": "Directed links after resolve",
                "n": len(edges),
            },
            {"item": "Self-links skipped", "n": self_links},
            {"item": "Reciprocal pairs", "n": len(reciprocal_pairs)},
            {"item": "Missing reverse (listed)", "n": len(missing)},
            {"item": "Unresolved partner mentions", "n": len(unresolved)},
            {
                "item": "Distinct unresolved partner spellings",
                "n": (
                    unresolved["partner_raw"].str.casefold().nunique()
                    if not unresolved.empty
                    else 0
                ),
            },
            {"item": "Partner resolved but no Sheet2 row", "n": len(no_row)},
        ]
    )

    readme = pd.DataFrame(
        {
            "Sheet2 missing reverse list": [
                "One row = one missing reverse after homeland resolve.",
                "",
                "has_link: Sheet2 group that already lists the partner.",
                "wrote: the partner spelling as coded in Joking link.",
                "missing_on: Sheet2 row(s) that should name them back.",
                "should_add: name to add on missing_on's Joking link "
                "(or remove the partner from has_link with a note).",
                "",
                "unresolved_partners: spellings that did not resolve — fix aliases "
                "first; these are not counted as missing reverse.",
                "partner_not_in_sheet2: partner resolved, but there is no Sheet2 row.",
                "",
                "Resolve: ethnic_entity_index → registry aliases → GIS exact.",
            ]
        }
    )

    return {
        "README": readme,
        "missing_reverse": missing,
        "unresolved_partners": unresolved,
        "partner_not_in_sheet2": no_row,
        "summary": summary,
    }


def _style_sheet(ws) -> None:
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    headers = [c.value for c in ws[1]]
    widths = {
        "has_link": 18,
        "wrote": 22,
        "resolved_as": 22,
        "missing_on": 18,
        "should_add": 16,
        "missing_on_currently_lists": 48,
        "missing_on_unresolved_spellings": 28,
        "partner_raw": 20,
        "named_by": 16,
        "why": 48,
        "note": 40,
        "item": 48,
        "n": 10,
        "Region": 14,
        "Country_1": 14,
    }
    for i, name in enumerate(headers, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = widths.get(name, 14)

    wrap_names = {
        "missing_on_currently_lists",
        "missing_on_unresolved_spellings",
        "why",
        "note",
        "wrote",
    }
    wrap_idx = {i for i, name in enumerate(headers, start=1) if name in wrap_names}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap if cell.column in wrap_idx else top

    ws.row_dimensions[1].height = 22


def write_xlsx(sheets: dict[str, pd.DataFrame], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    order = [
        "README",
        "missing_reverse",
        "unresolved_partners",
        "partner_not_in_sheet2",
        "summary",
    ]
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name in order:
            df = sheets.get(name)
            if df is None:
                continue
            df.to_excel(writer, sheet_name=name[:31], index=False)
        for ws in writer.book.worksheets:
            _style_sheet(ws)


def run(src: Path) -> dict[str, pd.DataFrame]:
    ensure_output_dirs()
    print(f"Loading Sheet2: {src}")
    df = load_sheet2(src)
    print(f"  rows={len(df)}")
    print("Resolving homelands…")
    cache = ResolveCache(resolver=get_resolver())
    sheets = build_tables(df, cache)
    n_miss = len(sheets["missing_reverse"])
    n_unres = len(sheets["unresolved_partners"])
    n_norow = len(sheets["partner_not_in_sheet2"])
    print(
        f"  missing_reverse={n_miss}  "
        f"unresolved_partners={n_unres}  "
        f"partner_not_in_sheet2={n_norow}"
    )
    return sheets


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Sheet2 one-sided tables (use export_ra_workpack for the RA file)"
    )
    parser.add_argument("--src", type=Path, default=ICMID_MANUAL_XLSX)
    args = parser.parse_args()
    run(args.src)
    print("Internal helper — run: uv run python -B code/jr_database/export_ra_workpack.py")


if __name__ == "__main__":
    main()
