#!/usr/bin/env python3
"""ICMID Sheet2/Sheet3 import helpers for export_ra_workpack (internal).

Do not run directly — use export_ra_workpack.py.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_JR = Path(__file__).resolve().parent.parent  # code/jr_database
_CODE = _JR.parent
_PIPELINE = _CODE.parent
for _p in (_CODE, _PIPELINE, _CODE / "visualization"):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

from jr_database.config import (  # noqa: E402
    ICMID_MANUAL_XLSX,
    ensure_output_dirs,
)
from jr_database.resolve_homeland import HomelandHit, get_resolver  # noqa: E402
from jr_database.sources import (  # noqa: E402
    _clean,
    _parse_joking_partners,
    load_all_cross_assertions,
)

FOCUS_SOURCES = frozenset({"keerthana_analysis", "llm_ehraf"})
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _excel_safe(val: Any) -> Any:
    if isinstance(val, str):
        return _ILLEGAL_XLSX_RE.sub("", val)
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return val


def _partner_key(hit: HomelandHit, raw: str) -> str:
    if hit.polygon_source in {"murdock", "greg", "geopr"} and hit.polygon_id:
        return f"{hit.polygon_source}:{hit.polygon_id.upper()}"
    return f"raw:{_clean(raw).casefold()}"


def _load_sheet2(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Sheet2")
    df.columns = [str(c).strip() for c in df.columns]
    df["Ethnic Group"] = df["Ethnic Group"].map(
        lambda v: re.sub(r"\s+", " ", _clean(v))
    )
    return df


def build_worklist(src: Path) -> dict[str, pd.DataFrame]:
    resolver = get_resolver()
    sheet2 = _load_sheet2(src)

    # Sheet2 murdock_pid → Ethnic Group name(s) + current outgoing partner keys
    pid_to_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    outgoing: dict[str, set[str]] = defaultdict(set)

    for i, row in sheet2.iterrows():
        name = _clean(row.get("Ethnic Group"))
        if not name:
            continue
        hit = resolver.resolve(name)
        if hit.polygon_source != "murdock" or not hit.polygon_id:
            continue
        pid = hit.polygon_id.upper()
        jl = row.get("Joking link") if "Joking link" in sheet2.columns else ""
        pid_to_rows[pid].append(
            {
                "excel_row": int(i) + 2,  # 1-based + header
                "Ethnic Group": name,
                "Region": _clean(row.get("Region")),
                "Country_1": _clean(row.get("Country_1")),
                "Joking link": _excel_safe(jl) if _clean(jl) else "",
            }
        )
        for partner in _parse_joking_partners(jl):
            ph = resolver.resolve(partner)
            outgoing[pid].add(_partner_key(ph, partner))

    assertions = load_all_cross_assertions()
    focus = assertions[assertions["source_dataset"].isin(FOCUS_SOURCES)].copy()

    # Aggregate unique undirected pairs
    pairs: dict[tuple[str, str], dict[str, Any]] = {}
    for _, r in focus.iterrows():
        a = _clean(r.get("entity_a"))
        b = _clean(r.get("entity_b"))
        if not a or not b or a.casefold() == b.casefold():
            continue
        pk = tuple(sorted([a.casefold(), b.casefold()]))
        if pk not in pairs:
            pairs[pk] = {
                "entity_a": a if a.casefold() <= b.casefold() else b,
                "entity_b": b if a.casefold() <= b.casefold() else a,
                "sources": set(),
                "quotes": [],
                "types_a": set(),
                "types_b": set(),
                "regions": set(),
            }
        rec = pairs[pk]
        # keep original orientation labels loosely via sorted names already
        rec["sources"].add(_clean(r.get("source_dataset")))
        q = _clean(r.get("quote"))
        if q and q not in rec["quotes"]:
            rec["quotes"].append(q)
        ta = _clean(r.get("entity_a_type"))
        tb = _clean(r.get("entity_b_type"))
        # map types onto sorted entity labels
        if a.casefold() <= b.casefold():
            if ta:
                rec["types_a"].add(ta)
            if tb:
                rec["types_b"].add(tb)
        else:
            if ta:
                rec["types_b"].add(ta)
            if tb:
                rec["types_a"].add(tb)
        reg = _clean(r.get("region"))
        if reg:
            rec["regions"].add(reg)

    to_sheet2: list[dict[str, Any]] = []
    to_sheet3: list[dict[str, Any]] = []
    already_ok: list[dict[str, Any]] = []

    for rec in pairs.values():
        a, b = rec["entity_a"], rec["entity_b"]
        ha, hb = resolver.resolve(a), resolver.resolve(b)
        a_m = ha.polygon_source == "murdock" and bool(ha.polygon_id)
        b_m = hb.polygon_source == "murdock" and bool(hb.polygon_id)
        sources = ";".join(sorted(rec["sources"]))
        region = "; ".join(sorted(rec["regions"]))
        quote = _excel_safe(" | ".join(rec["quotes"])[:1500])

        if not a_m and not b_m:
            to_sheet3.append(
                {
                    "entity_a": a,
                    "entity_a_type": "; ".join(sorted(rec["types_a"])),
                    "entity_a_resolve": ha.polygon_source or "(none)",
                    "entity_a_polygon_id": ha.polygon_id,
                    "entity_b": b,
                    "entity_b_type": "; ".join(sorted(rec["types_b"])),
                    "entity_b_resolve": hb.polygon_source or "(none)",
                    "entity_b_polygon_id": hb.polygon_id,
                    "source_flags": sources,
                    "region": region,
                    "quote": quote,
                    "action": "add to Sheet3 (neither side is murdock)",
                }
            )
            continue

        # At least one murdock — emit directed Sheet2 actions for each murdock side
        murdock_sides: list[tuple[str, HomelandHit, str, HomelandHit]] = []
        if a_m:
            murdock_sides.append((a, ha, b, hb))
        if b_m:
            murdock_sides.append((b, hb, a, ha))

        for m_name, m_hit, p_name, p_hit in murdock_sides:
            pid = m_hit.polygon_id.upper()
            rows = pid_to_rows.get(pid, [])
            sheet2_exists = bool(rows)
            sheet2_names = "; ".join(r["Ethnic Group"] for r in rows) if rows else ""
            sheet2_excel_rows = "; ".join(str(r["excel_row"]) for r in rows) if rows else ""
            current_jl = " | ".join(
                f"{r['Ethnic Group']}: {r['Joking link'] or '(empty)'}" for r in rows
            )
            pkey = _partner_key(p_hit, p_name)
            already = pkey in outgoing.get(pid, set())

            if sheet2_exists and already:
                status = "already_on_sheet2"
                action = "skip — partner already on this Murdock row"
            elif sheet2_exists:
                status = "add_partner_to_existing_row"
                action = f"add '{p_name}' to Joking link on Sheet2 row(s) {sheet2_excel_rows}"
            else:
                status = "create_new_sheet2_row"
                action = (
                    f"create Sheet2 row Ethnic Group={pid} (or {m_name}), "
                    f"then set Joking link to include '{p_name}'"
                )

            out = {
                "status": status,
                "action": action,
                "murdock_polygon_id": pid,
                "murdock_side_name": m_name,
                "sheet2_Ethnic_Group": sheet2_names,
                "sheet2_excel_row": sheet2_excel_rows,
                "partner_to_add": p_name,
                "partner_resolve": p_hit.polygon_source or "(none)",
                "partner_polygon_id": p_hit.polygon_id,
                "partner_type": "",
                "source_flags": sources,
                "region": region,
                "current_Joking_link": _excel_safe(current_jl[:800]),
                "quote": quote,
            }
            # attach partner type if known
            if p_name.casefold() == a.casefold():
                out["partner_type"] = "; ".join(sorted(rec["types_a"]))
            elif p_name.casefold() == b.casefold():
                out["partner_type"] = "; ".join(sorted(rec["types_b"]))

            if status == "already_on_sheet2":
                already_ok.append(out)
            else:
                to_sheet2.append(out)

    status_rank = {
        "create_new_sheet2_row": 0,
        "add_partner_to_existing_row": 1,
        "already_on_sheet2": 2,
    }
    sheet2_df = pd.DataFrame(to_sheet2)
    if not sheet2_df.empty:
        sheet2_df["_rk"] = sheet2_df["status"].map(status_rank)
        sheet2_df = (
            sheet2_df.sort_values(
                ["_rk", "murdock_polygon_id", "partner_to_add"], kind="mergesort"
            )
            .drop(columns=["_rk"])
            .reset_index(drop=True)
        )

    sheet3_df = pd.DataFrame(to_sheet3)
    if not sheet3_df.empty:
        sheet3_df = sheet3_df.sort_values(
            ["entity_a", "entity_b"], kind="mergesort"
        ).reset_index(drop=True)

    already_df = pd.DataFrame(already_ok)
    if not already_df.empty:
        already_df = already_df.sort_values(
            ["murdock_polygon_id", "partner_to_add"], kind="mergesort"
        ).reset_index(drop=True)

    n_add = int((sheet2_df["status"] == "add_partner_to_existing_row").sum()) if len(sheet2_df) else 0
    n_new = int((sheet2_df["status"] == "create_new_sheet2_row").sum()) if len(sheet2_df) else 0

    summary = pd.DataFrame(
        [
            {"item": "Focus sources", "n": "keerthana_analysis; llm_ehraf"},
            {"item": "Unique focus pairs", "n": len(pairs)},
            {"item": "to_sheet2 — add partner on existing row", "n": n_add},
            {"item": "to_sheet2 — create new Murdock row", "n": n_new},
            {"item": "to_sheet2 total (needs action)", "n": len(sheet2_df)},
            {"item": "already on Sheet2 (skip)", "n": len(already_df)},
            {"item": "to_sheet3 — neither side murdock", "n": len(sheet3_df)},
        ]
    )

    readme = pd.DataFrame(
        {
            "ICMID Sheet2 / Sheet3 import worklist": [
                "Advisor rule: ≥1 side resolves to murdock → Sheet2; else → Sheet3.",
                "",
                "Sources: keerthana_analysis + llm_ehraf (ICMID sheet pairs not used as input).",
                "",
                "Sheet to_sheet2 — rows that still need coding on ICMID Sheet2.",
                "  status=add_partner_to_existing_row → append partner_to_add onto that row's Joking link.",
                "  status=create_new_sheet2_row → Sheet2 has no Murdock row yet; create Ethnic Group then add partner.",
                "  sheet2_excel_row is the Excel row number in ICMID- Africa.xlsx Sheet2.",
                "",
                "Sheet to_sheet3 — neither endpoint is murdock; candidate lines for Sheet3.",
                "",
                "Sheet already_on_sheet2 — already present; no action.",
                "",
                "Review before merge: drop kin_role / villagers / overly vague partners.",
                f"Coding file: {src.name}",
            ]
        }
    )

    return {
        "README": readme,
        "to_sheet2": sheet2_df,
        "to_sheet3": sheet3_df,
        "already_on_sheet2": already_df,
        "summary": summary,
    }


def _style(ws) -> None:
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    fill_new = PatternFill("solid", fgColor="F8CBAD")
    fill_add = PatternFill("solid", fgColor="FFF2CC")
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    headers = [c.value for c in ws[1]]
    widths = {
        "status": 28,
        "action": 48,
        "murdock_polygon_id": 16,
        "murdock_side_name": 18,
        "sheet2_Ethnic_Group": 18,
        "partner_to_add": 22,
        "current_Joking_link": 40,
        "quote": 40,
        "entity_a": 22,
        "entity_b": 22,
        "item": 48,
        "n": 36,
    }
    for i, name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 14)

    status_col = headers.index("status") + 1 if "status" in headers else None
    wrap_names = {"action", "quote", "current_Joking_link", "notes"}
    wrap_idx = {i for i, n in enumerate(headers, start=1) if n in wrap_names}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap if cell.column in wrap_idx else top
        if status_col:
            val = row[status_col - 1].value
            if val == "create_new_sheet2_row":
                for cell in row:
                    cell.fill = fill_new
            elif val == "add_partner_to_existing_row":
                for cell in row:
                    cell.fill = fill_add


def write_xlsx(sheets: dict[str, pd.DataFrame], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    order = ["README", "to_sheet2", "to_sheet3", "already_on_sheet2", "summary"]
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name in order:
            df = sheets.get(name)
            if df is None:
                continue
            df.to_excel(writer, sheet_name=name[:31], index=False)
        for ws in writer.book.worksheets:
            _style(ws)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="ICMID Sheet2/Sheet3 import tables (use export_ra_workpack for the RA file)"
    )
    parser.add_argument("--src", type=Path, default=ICMID_MANUAL_XLSX)
    args = parser.parse_args()

    ensure_output_dirs()
    print(f"Building worklist from {args.src.name} + keerthana_analysis/llm_ehraf…")
    sheets = build_worklist(args.src)
    s2 = sheets["to_sheet2"]
    n_add = int((s2["status"] == "add_partner_to_existing_row").sum()) if len(s2) else 0
    n_new = int((s2["status"] == "create_new_sheet2_row").sum()) if len(s2) else 0
    print(
        f"  to_sheet2: {len(s2)} "
        f"(add_partner={n_add}, new_row={n_new})  "
        f"to_sheet3={len(sheets['to_sheet3'])}  "
        f"already_ok={len(sheets['already_on_sheet2'])}"
    )
    print("Internal helper — run: uv run python -B code/jr_database/export_ra_workpack.py")


if __name__ == "__main__":
    main()
