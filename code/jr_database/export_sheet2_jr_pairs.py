#!/usr/bin/env python3
"""Expand ICMID Sheet2 into a JR-pairs sheet (one row = one undirected pair).

Dedupes by resolved homeland polygon when possible (aliases like
Kugni/KUNYI or Wasili/SHUWA collapse to one pair). Source fields prefer
the Sheet2 side whose ``Source_Quote`` mentions the partner **or any of
its aliases**. If neither quote matches, prefer the side with
``Other sources`` filled and flag ``source_review``.

Audited rows (Auditor filled — currently East Africa) are preserved.
Only exact identical duplicates within audited rows are dropped.

Usage (from ICMID PingJu project root):
    uv run python -B code/jr_database/export_sheet2_jr_pairs.py
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

_CODE = Path(__file__).resolve().parent.parent
_PIPELINE = _CODE.parent
for _p in (_CODE, _PIPELINE, _CODE / "visualization"):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

from entity_index import compute_homeland_found, load_index  # noqa: E402
from jr_database.config import ICMID_MANUAL_XLSX, RESULT_OUTPUT, ensure_output_dirs  # noqa: E402
from jr_database.resolve_homeland import HomelandResolver, get_resolver  # noqa: E402
from jr_database.sources import _clean, _parse_joking_partners  # noqa: E402

PAIR_SHEET = "JR_pair"
PAIR_SHEET_ALIASES = ("JR_pair", "JR pair", "Sheet2_JR_pairs")
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

PAIR_COLS = (
    "entity_a",
    "entity_b",
    "Region",
    "Sources",
    "Source_Full_Citation",
    "Source_Page",
    "Source_Quote",
    "Source_File",
    "Source_URL",
    "Other sources",
    "Notes",
    "Coder",
    "source_review",
    "Auditor",
    "Date",
)

SOURCE_COLS = (
    "Region",
    "Sources",
    "Source_Full_Citation",
    "Source_Page",
    "Source_Quote",
    "Source_File",
    "Source_URL",
    "Other sources",
    "Notes",
    "Coder",
)

REVIEW_OK = ""
REVIEW_NEEDS = (
    "NEEDS_REVIEW: Source_Quote mentions neither partner nor aliases; "
    "prefer Other sources when present"
)


def _excel_safe(val: Any) -> Any:
    if isinstance(val, str):
        return _ILLEGAL_XLSX_RE.sub("", val)
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return val


def _name_in_text(text: str, name: str) -> bool:
    """Case-insensitive mention; short names use word boundaries."""
    t = _clean(text)
    n = _clean(name)
    if not t or not n:
        return False
    tl, nl = t.casefold(), n.casefold()
    if len(nl) <= 3:
        return bool(re.search(rf"\b{re.escape(nl)}\b", tl))
    return nl in tl


def _poly_key(resolver: HomelandResolver, name: str) -> str:
    hit = resolver.resolve(name)
    if hit.polygon_source and hit.polygon_id:
        return f"{hit.polygon_source}:{hit.polygon_id}".upper()
    return f"RAW:{_clean(name).casefold()}"


def _build_alias_sets(resolver: HomelandResolver) -> dict[str, set[str]]:
    """polygon_key → spellings known for that homeland (index + raw fallbacks)."""
    aliases: dict[str, set[str]] = defaultdict(set)
    index = load_index()
    for _, row in index.iterrows():
        raw = _clean(row.get("raw_value"))
        canon = _clean(row.get("canonical_name"))
        if compute_homeland_found(row):
            src = _clean(row.get("polygon_source")).lower().replace("geoepr", "geopr")
            pid = _clean(row.get("polygon_id")).upper()
            key = f"{src}:{pid}".upper()
        else:
            key = _poly_key(resolver, raw or canon)
        for name in (raw, canon):
            if name:
                aliases[key].add(name)
                # outer token without parenthetical: "Hausa (Katsina)" → "Hausa"
                outer = re.sub(r"\s*\(.*?\)\s*", " ", name).strip()
                if outer and outer.casefold() != name.casefold():
                    aliases[key].add(outer)
    return aliases


def _names_for(alias_sets: dict[str, set[str]], poly: str, fallback: str) -> set[str]:
    out = set(alias_sets.get(poly, set()))
    if fallback:
        out.add(fallback)
        outer = re.sub(r"\s*\(.*?\)\s*", " ", fallback).strip()
        if outer:
            out.add(outer)
    return {n for n in out if n}


def _quote_mentions_any(quote: str, names: set[str]) -> bool:
    return any(_name_in_text(quote, n) for n in names)


def _pick_source_candidate(
    candidates: list[dict[str, Any]],
    alias_sets: dict[str, set[str]],
) -> tuple[dict[str, str], str]:
    """Prefer quote that names partner (or aliases); else prefer Other sources."""
    if not candidates:
        empty = {c: "" for c in SOURCE_COLS}
        return empty, REVIEW_NEEDS

    scored: list[tuple[tuple[int, int, int], int, dict[str, Any]]] = []
    for i, cand in enumerate(candidates):
        quote = cand.get("Source_Quote", "")
        partner_names = _names_for(
            alias_sets, cand.get("partner_poly", ""), cand.get("partner", "")
        )
        named_names = _names_for(
            alias_sets, cand.get("named_poly", ""), cand.get("named_by", "")
        )
        mentions_partner = 1 if _quote_mentions_any(quote, partner_names) else 0
        mentions_self = 1 if _quote_mentions_any(quote, named_names) else 0
        has_other = 1 if _clean(cand.get("Other sources", "")) else 0
        # primary: partner/alias in quote; then Other sources; then self mention
        scored.append(((mentions_partner, has_other, mentions_self), i, cand))

    scored.sort(key=lambda t: (-t[0][0], -t[0][1], -t[0][2], t[1]))
    best_key, _, best = scored[0]
    fields = {c: _excel_safe(best.get(c, "")) for c in SOURCE_COLS}
    if best_key[0] >= 1:
        return fields, REVIEW_OK
    return fields, REVIEW_NEEDS


def _display_pair(
    left_poly: str,
    right_poly: str,
    candidates: list[dict[str, Any]],
) -> tuple[str, str]:
    """Pick stable display names for the undirected pair."""
    left_names: list[str] = []
    right_names: list[str] = []
    for c in candidates:
        if c.get("named_poly") == left_poly:
            left_names.append(c["named_by"])
        if c.get("partner_poly") == left_poly:
            left_names.append(c["partner"])
        if c.get("named_poly") == right_poly:
            right_names.append(c["named_by"])
        if c.get("partner_poly") == right_poly:
            right_names.append(c["partner"])

    def pick(names: list[str], poly: str) -> str:
        if not names:
            return poly.split(":", 1)[-1] if ":" in poly else poly
        # prefer shortest UPPER-ish Sheet2 ethnic-group style (often murdock id)
        return sorted(names, key=lambda n: (len(n), n.casefold()))[0]

    a, b = pick(left_names, left_poly), pick(right_names, right_poly)
    if a.casefold() <= b.casefold():
        return a, b
    return b, a


def build_jr_pairs(
    sheet2: pd.DataFrame,
    resolver: HomelandResolver | None = None,
) -> pd.DataFrame:
    """One undirected JR pair per unique {homeland_a, homeland_b} (alias-aware)."""
    resolver = resolver or get_resolver()
    alias_sets = _build_alias_sets(resolver)

    df = sheet2.copy()
    df.columns = [str(c).strip() for c in df.columns]
    f_col = "Joking link" if "Joking link" in df.columns else df.columns[5]

    buckets: dict[tuple[str, str], dict[str, Any]] = {}

    for _, r in df.iterrows():
        a = _clean(r.get("Ethnic Group"))
        if not a:
            continue
        partners = _parse_joking_partners(r.get(f_col))
        if not partners:
            continue
        a_poly = _poly_key(resolver, a)
        carry = {
            c: _excel_safe(_clean(r.get(c)) if c in df.columns else "")
            for c in SOURCE_COLS
        }
        for b in partners:
            if not b or b.casefold() == a.casefold():
                continue
            b_poly = _poly_key(resolver, b)
            if a_poly == b_poly:
                continue
            left_poly, right_poly = sorted([a_poly, b_poly])
            key = (left_poly, right_poly)
            bucket = buckets.get(key)
            if bucket is None:
                bucket = {
                    "left_poly": left_poly,
                    "right_poly": right_poly,
                    "candidates": [],
                }
                buckets[key] = bucket
            bucket["candidates"].append(
                {
                    "named_by": a,
                    "partner": b,
                    "named_poly": a_poly,
                    "partner_poly": b_poly,
                    **carry,
                }
            )

    rows: list[dict[str, Any]] = []
    for bucket in buckets.values():
        fields, review = _pick_source_candidate(bucket["candidates"], alias_sets)
        entity_a, entity_b = _display_pair(
            bucket["left_poly"], bucket["right_poly"], bucket["candidates"]
        )
        rows.append(
            {
                "entity_a": entity_a,
                "entity_b": entity_b,
                **fields,
                "source_review": review,
                "Auditor": "",
                "Date": "",
                "_pair_key": (bucket["left_poly"], bucket["right_poly"]),
            }
        )

    cols = list(PAIR_COLS) + ["_pair_key"]
    if not rows:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame(rows)[cols]
    return out.sort_values(["entity_a", "entity_b"], kind="mergesort").reset_index(drop=True)


def _is_east_region(reg: Any) -> bool:
    return "east" in str(reg or "").casefold()


def _row_pair_key(resolver: HomelandResolver, a: Any, b: Any) -> tuple[str, str]:
    return tuple(sorted([_poly_key(resolver, str(a)), _poly_key(resolver, str(b))]))


def load_existing_pairs(path: Path, sheet_name: str) -> pd.DataFrame:
    if not path.is_file():
        return pd.DataFrame(columns=list(PAIR_COLS))
    names = [sheet_name, *PAIR_SHEET_ALIASES]
    seen: set[str] = set()
    df = None
    for name in names:
        if name in seen:
            continue
        seen.add(name)
        try:
            df = pd.read_excel(path, sheet_name=name)
            break
        except ValueError:
            continue
    if df is None:
        return pd.DataFrame(columns=list(PAIR_COLS))
    df.columns = [str(c).strip() for c in df.columns]
    for c in PAIR_COLS:
        if c not in df.columns:
            df[c] = ""
    return df[list(PAIR_COLS)]


def preserve_audited(
    existing: pd.DataFrame,
    resolver: HomelandResolver,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Keep Auditor-filled rows; collapse exact + alias dups.

    Alias dups keep the row where both spellings are Murdock homeland names.
    Returns (preserved_unique, empty_report).
    """
    if existing.empty:
        empty = pd.DataFrame(columns=list(PAIR_COLS) + ["_pair_key"])
        return empty, empty

    audited = existing[existing["Auditor"].map(lambda v: bool(_clean(v)))].copy()
    if audited.empty:
        empty = pd.DataFrame(columns=list(PAIR_COLS) + ["_pair_key"])
        return empty, empty

    audited["_pair_key"] = [
        _row_pair_key(resolver, a, b)
        for a, b in zip(audited["entity_a"], audited["entity_b"])
    ]
    audited["_name_key"] = [
        tuple(sorted([_clean(a).casefold(), _clean(b).casefold()]))
        for a, b in zip(audited["entity_a"], audited["entity_b"])
    ]

    def richness(row: pd.Series) -> int:
        return sum(1 for c in SOURCE_COLS if _clean(row.get(c)))

    audited["_rich"] = audited.apply(richness, axis=1)
    audited = audited.sort_values(["_rich"], ascending=False, kind="mergesort")
    preserved = audited.drop_duplicates(subset=["_name_key"], keep="first").copy()

    # Alias-level: keep Murdock-named spelling on both sides when possible
    preserved = dedupe_prefer_murdock_names(preserved, resolver)
    out = preserved[list(PAIR_COLS) + ["_pair_key"]].copy()
    empty_report = pd.DataFrame(columns=list(PAIR_COLS) + ["_pair_key"])
    return out.reset_index(drop=True), empty_report


def dedupe_prefer_murdock_names(
    df: pd.DataFrame,
    resolver: HomelandResolver,
) -> pd.DataFrame:
    """Within same polygon-pair, keep the row whose names are Murdock ids."""
    if df.empty:
        return df

    work = df.copy()
    if "_pair_key" not in work.columns:
        work["_pair_key"] = [
            _row_pair_key(resolver, a, b)
            for a, b in zip(work["entity_a"], work["entity_b"])
        ]

    def side_murdock_match(entity: Any) -> tuple[int, int]:
        hit = resolver.resolve(str(entity))
        if not hit.polygon_source or not hit.polygon_id:
            return (0, 0)
        is_murdock = 1 if hit.polygon_source.lower() == "murdock" else 0
        name_ok = 1 if _clean(entity).casefold() == hit.polygon_id.casefold() else 0
        return (is_murdock, 1 if (is_murdock and name_ok) else 0)

    def row_score(row: pd.Series) -> tuple:
        a_m, a_ok = side_murdock_match(row["entity_a"])
        b_m, b_ok = side_murdock_match(row["entity_b"])
        both_murdock_names = 1 if (a_ok and b_ok) else 0
        murdock_name_sides = a_ok + b_ok
        both_murdock_polys = 1 if (a_m and b_m) else 0
        audited = 1 if _clean(row.get("Auditor")) else 0
        rich = sum(1 for c in SOURCE_COLS if _clean(row.get(c)))
        return (both_murdock_names, murdock_name_sides, both_murdock_polys, audited, rich)

    keep_idx: list[Any] = []
    for _, group in work.groupby("_pair_key", sort=False):
        if len(group) == 1:
            keep_idx.append(group.index[0])
            continue
        ranked = sorted(group.index, key=lambda i: row_score(work.loc[i]), reverse=True)
        keep_idx.append(ranked[0])

    return work.loc[keep_idx].copy()


def merge_preserved_and_built(
    preserved: pd.DataFrame,
    built: pd.DataFrame,
) -> pd.DataFrame:
    """Audited rows win; drop rebuilt pairs that collide on polygon key."""
    keep_keys = set(preserved["_pair_key"].tolist()) if len(preserved) else set()
    fresh = built[~built["_pair_key"].isin(keep_keys)].copy() if len(built) else built
    parts = [df for df in (preserved, fresh) if len(df)]
    if not parts:
        return pd.DataFrame(columns=list(PAIR_COLS))
    out = pd.concat(parts, ignore_index=True)
    out = out[list(PAIR_COLS)]
    return out.sort_values(["entity_a", "entity_b"], kind="mergesort").reset_index(drop=True)


def _style_ws(ws) -> None:
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")
    review_fill = PatternFill("solid", fgColor="FCE4D6")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    headers = [c.value for c in ws[1]]
    wide = {
        "entity_a": 16,
        "entity_b": 16,
        "Region": 16,
        "Sources": 28,
        "Source_Full_Citation": 48,
        "Source_Page": 12,
        "Source_Quote": 48,
        "Source_File": 28,
        "Source_URL": 36,
        "Other sources": 28,
        "Notes": 36,
        "Coder": 16,
        "source_review": 40,
        "Auditor": 12,
        "Date": 12,
    }
    for i, name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wide.get(str(name), 14)

    wrap_names = {
        "Sources",
        "Source_Full_Citation",
        "Source_Quote",
        "Source_File",
        "Source_URL",
        "Other sources",
        "Notes",
        "source_review",
    }
    wrap_idx = {i for i, n in enumerate(headers, start=1) if n in wrap_names}
    review_col = headers.index("source_review") + 1 if "source_review" in headers else None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap if cell.column in wrap_idx else top
        if review_col:
            val = row[review_col - 1].value
            if val:
                for cell in row:
                    cell.fill = review_fill


def write_pairs_into_workbook(pairs: pd.DataFrame, path: Path, sheet_name: str = PAIR_SHEET) -> None:
    wb = load_workbook(path)
    # Replace canonical sheet and drop all legacy names (incl. spaced "JR pair").
    for name in {sheet_name, *PAIR_SHEET_ALIASES}:
        if name in wb.sheetnames:
            del wb[name]
    ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(pairs, index=False, header=True):
        ws.append([_excel_safe(v) for v in row])
    _style_ws(ws)
    wb.save(path)


def write_pairs_standalone(pairs: pd.DataFrame, path: Path, sheet_name: str = PAIR_SHEET) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pairs.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        _style_ws(writer.book[sheet_name[:31]])


def main() -> None:
    parser = argparse.ArgumentParser(description="Expand ICMID Sheet2 into JR pair rows")
    parser.add_argument("--src", type=Path, default=ICMID_MANUAL_XLSX)
    parser.add_argument("--sheet-name", default=PAIR_SHEET)
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Optional standalone xlsx copy (canonical sheet is JR_pair on ICMID workbook)",
    )
    parser.add_argument(
        "--no-write-src",
        action="store_true",
        help="Do not modify the ICMID workbook (requires --out)",
    )
    parser.add_argument(
        "--rebuild-all",
        action="store_true",
        help="Ignore audited rows and rebuild everything from Sheet2",
    )
    parser.add_argument(
        "--write-alias-report",
        action="store_true",
        help="Also write jr_pair_east_africa_alias_dups.xlsx when audited alias dups exist",
    )
    args = parser.parse_args()
    ensure_output_dirs()

    if not args.src.is_file():
        raise SystemExit(f"Missing workbook: {args.src}")
    if args.no_write_src and args.out is None:
        raise SystemExit("--no-write-src requires --out")

    get_resolver.cache_clear()
    resolver = get_resolver()

    print(f"Loading Sheet2 from {args.src}")
    sheet2 = pd.read_excel(args.src, sheet_name="Sheet2")
    built = build_jr_pairs(sheet2, resolver=resolver)
    print(f"  rebuilt from Sheet2 (alias-deduped): {len(built)} pairs")

    existing = load_existing_pairs(args.src, args.sheet_name)
    if args.rebuild_all:
        pairs = built[list(PAIR_COLS)]
        east_report = pd.DataFrame()
        print("  --rebuild-all: not preserving audited rows")
    else:
        preserved, east_report = preserve_audited(existing, resolver)
        print(f"  preserved audited (Murdock-name preferred on alias dups): {len(preserved)}")
        pairs = merge_preserved_and_built(preserved, built)

    n_review = int((pairs["source_review"].fillna("").astype(str).str.len() > 0).sum())
    n_east = int(pairs["Region"].map(_is_east_region).sum())
    n_audited = int(pairs["Auditor"].map(lambda v: bool(_clean(v))).sum())
    print(
        f"  final JR_pair rows={len(pairs)}  "
        f"east={n_east}  audited={n_audited}  needs_source_review={n_review}"
    )

    if not args.no_write_src:
        write_pairs_into_workbook(pairs, args.src, sheet_name=args.sheet_name)
        print(f"  → sheet {args.sheet_name!r} on {args.src}")

    if args.out is not None:
        write_pairs_standalone(pairs, args.out, sheet_name=args.sheet_name)
        print(f"  → {args.out}")

    # Remove stale standalone copy if we stopped writing it by default
    stale = RESULT_OUTPUT / "sheet2_jr_pairs.xlsx"
    if args.out is None and stale.is_file():
        stale.unlink()
        print(f"  removed duplicate {stale.name} (canonical: JR_pair sheet on ICMID workbook)")

    if args.write_alias_report and len(east_report):
        report_path = RESULT_OUTPUT / "jr_pair_east_africa_alias_dups.xlsx"
        cols = [c for c in PAIR_COLS if c in east_report.columns] + ["_pair_key"]
        east_report[cols].to_excel(report_path, index=False)
        print(f"  → East Africa alias-dup report: {report_path}")


if __name__ == "__main__":
    main()
